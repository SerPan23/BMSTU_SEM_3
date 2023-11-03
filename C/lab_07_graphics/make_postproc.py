import math
import matplotlib.pyplot as plt
import collections
from openpyxl import Workbook

STYLES = [
    ('lightsteelblue', 's'),
    ('royalblue', 'o'),
    ('tomato', 'x'),
    ('limegreen', 'h'),
    ('hotpink', 'v'),
    ('orange', '^'),
]

PLOT_SIZE = (20, 10)
DPI=1000

def collect_data_for_box_plot(filename):
    entries = collections.defaultdict(list)
    entries = dict()
    with open(filename, 'r') as f:
        for line in f:
            its = line.split()
            if not its:
                continue

            size, value = its
            # entries[int(size)].append(float(value))
            if entries.get(int(size)) != None:
                entries[int(size)].append(float(value))
            else:
                entries[int(size)] = [float(value)]

    return entries

def make_data_for_box_plot(init_type):
    with open(f'./postproc_data/4/boxdata_{init_type}.txt', 'w') as graph_4: 
        for size in sizes:
            f = open(f'./data/main_index_O2_{init_type}_{size}.txt')
            # params = f.readlines()
            for line in f:
                graph_4.write(f'{size} {line}')
            f.close()

def get_x_y_data(graph_data):
    x = []
    y = []
    for i in graph_data:
        t = list(map(float, i.split()))
        x.append(t[0])
        y.append(t[1])
    
    return x, y

def get_error_data(graph_data):
    low = []
    high = []
    for i in graph_data:
        t = list(map(float, i.split()))
        high.append(t[2] - t[1])
        low.append(t[1] - t[3])
    
    return low, high


with open('config.ini') as config:
    names = config.readline().split()
    opts = config.readline().split()
    init_types = config.readline().split()
    sizes = config.readline().split()

for name in names:
    for opt in opts:
        for init_type in init_types:
            wb_path_save = ''
            if int(init_type) == 1:
                graph = open(f'./postproc_data/1/{name}_{opt}.txt', 'w')
                wb_path_save = f'./postproc_data/1/{name}_{opt}.xlsx'
            elif int(init_type) == 0:
                graph = open(f'./postproc_data/2/{name}_{opt}.txt', 'w')
                wb_path_save = f'./postproc_data/2/{name}_{opt}.xlsx'

            if opt == 'O2':
                graph_3 = open(f'./postproc_data/3/{name}_{init_type}.txt', 'w')
                wb_3_path_save = f'./postproc_data/3/{name}_{init_type}.xlsx'
                wb_3 = Workbook()
                ws_3 = wb.active
                ws_3.append(("Размер массива", "Время выполнения ", 'Величина относительной стандартной ошибки среднего'))

            if name == 'main_index' and opt == 'O2':
                graph_4 = open(f'./postproc_data/4/data_{init_type}.txt', 'w')
                wb_4_path_save = f'./postproc_data/4/data_{init_type}.xlsx'
                wb_4 = Workbook()
                ws_4 = wb.active
                ws_4.append(("Размер массива", "Время выполнения ", 'Величина относительной стандартной ошибки среднего'))

            wb = Workbook()
            ws = wb.active
            ws.append(("Размер массива", "Время выполнения ", 'Величина относительной стандартной ошибки среднего'))


            for size in sizes:
                params = open(f'./preproc_data/{name}_{opt}_{init_type}_{size}.txt').readlines()

                mean = params[0].strip()
                median = params[1].strip()
                maxx = params[2].strip()
                minn = params[3].strip()
                per_25 = params[4].strip()
                per_75 = params[5].strip()
                variance = float(params[6].strip())

                s = math.sqrt(variance)
                stderr = s/math.sqrt(int(size))
                try:
                    rse = stderr/float(mean) * 100
                except ZeroDivisionError:
                    rse = 0.0

                ws.append((size, mean, str(rse)))

                graph.write(f'{size} {mean}\n')

                if opt == 'O2':
                    graph_3.write(f'{size} {mean} {maxx} {minn}\n')
                    ws_3.append((size, mean, str(rse)))

                if name == 'main_index' and opt == 'O2':
                    graph_4.write(f'{size} {mean} {maxx} {minn} {per_25} {median} {per_75}\n')
                    ws_4.append((size, mean, str(rse)))

            
            wb.save(wb_path_save)
            wb.close()

            graph.close()
            try:
                wb_3.save(wb_3_path_save)
                wb_3.close()
                graph_3.close()
            except:
                pass
            try:
                wb_4.save(wb_4_path_save)
                wb_4.close()
                graph_4.close()
            except:
                pass

try:
    plt.figure(figsize=PLOT_SIZE)
    plt.grid(True)
    plt.xlabel("Размер массива")
    plt.ylabel("Время")
    i = 0
    for name in names:
        for opt in opts:
            graph = open(f'./postproc_data/1/{name}_{opt}.txt', 'r').readlines()
            x, y = get_x_y_data(graph)
            plt.xticks(x, sizes, rotation=45, ha='right')
            plt.plot(x, y, color=STYLES[i][0], marker=STYLES[i][1], label=f'{name} {opt}')
            i += 1
    plt.legend()
    plt.savefig('linear_1.svg', dpi=DPI)
    plt.close()
except:
    pass

try:
    plt.figure(figsize=PLOT_SIZE)
    plt.grid(True)
    plt.xlabel("Размер массива")
    plt.ylabel("Время")
    i = 0
    for name in names:
        for opt in opts:
            graph = open(f'./postproc_data/2/{name}_{opt}.txt', 'r').readlines()
            x, y = get_x_y_data(graph)
            plt.xticks(x, sizes, rotation=45, ha='right')
            plt.plot(x, y, color=STYLES[i][0], marker=STYLES[i][1], label=f'{name} {opt}')
            i += 1
    plt.legend()
    plt.savefig('linear_2.svg', dpi=DPI)
    plt.close()
except:
    pass

try:
    plt.figure(figsize=PLOT_SIZE)
    plt.grid(True)
    plt.xlabel("Размер массива")
    plt.ylabel("Время")
    i = 0
    for name in names:
        for init_type in init_types:
            graph = open(f'./postproc_data/3/{name}_{init_type}.txt', 'r').readlines()
            x, y = get_x_y_data(graph)
            plt.xticks(x, sizes, rotation=45, ha='right')
            low, high = get_error_data(graph)
            label2 = 'is sorted' if int(init_type) else 'is not sorted'
            plt.plot(x, y, color=STYLES[i][0], marker=STYLES[i][1], label=f'{name} {label2}')
            plt.errorbar(x, y, [low, high], fmt=STYLES[i][0], linewidth=2, capsize=6)
            i += 1
    plt.legend()
    plt.savefig('errors.svg', dpi=DPI)
    plt.close()
except:
    pass

try:
    plt.figure(figsize=PLOT_SIZE)
    plt.grid(True)
    plt.xlabel("Размер массива")
    plt.ylabel("Время")
    i = 0
    for init_type in init_types:
        graph = open(f'./postproc_data/4/data_{init_type}.txt', 'r').readlines()
        x, y = get_x_y_data(graph)
        label = 'is sorted' if int(init_type) else 'is not sorted'
        plt.xticks(x, sizes, rotation=45, ha='right')
        plt.plot(x, y, color=STYLES[i][0], marker=STYLES[i][1], label=label)
        
        make_data_for_box_plot(init_type)
        box_data = collect_data_for_box_plot(f'./postproc_data/4/boxdata_{init_type}.txt')
        bp = plt.boxplot(box_data.values(), positions=list(box_data.keys()), 
                         showfliers=False, showmeans=False, showcaps=True, 
                         widths=100, capwidths=70, meanline=False, whis=0)
        
        low, high = get_error_data(graph)
        plt.errorbar(x, y, [low, high], fmt=STYLES[i][0], linewidth=2, capsize=6)
        i += 1
    plt.legend()
    plt.savefig('moustache.svg', dpi=DPI)
    plt.close()
except Exception as e:
    print(e)
